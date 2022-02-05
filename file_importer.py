#File to hold structures related to importing data from the user-edited spreadsheet

#    Opin Suomea - a program to help people learn Finnish grammar
#    Copyright (C) 2022 Marc Perkins
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.
#    Contact Marc at the email address specified in config.printcontactinfo()

import os
import time
import re
import opinsuomea_utils as osu
from openpyxl import load_workbook
import config

datafilename = "OpinSuomea - Lause tiedot.xlsx"
#datafilename = "OpinSuomea - Lause tiedot - error test.xlsx"

appfolder = os.path.dirname(__file__)
datafolder = os.path.join(appfolder, config.jsonconfigdata["data_folder"])
datafile = os.path.join(datafolder, datafilename)


def openfile():
    if os.path.exists(datafile):
         print("Data file '{}' exists.".format(datafile))
    else:
        print("It appears that the datafile does not exist.  Should have been:")
        print(datafile)
        exit()
    try:
        wb = load_workbook(filename = datafile)
        print("Workbook opened")
    except:
        print("Looks like workbook loading failed.")
        exit()
    print("\nHere are all the sheets we saw in the file")
    for sheet in wb:
        print(sheet.title)
    print("\nThis list should include only unit worksheets:")
    unitsheets = []
    for sheet in wb:
        if sheet.title == "Template":
            continue
        if sheet.title == "Usage":
            continue
        if sheet.title == "Verbit":
            verbsheet = sheet
            #print(sheet.title)
            continue
        #Everything else should be unit sheets
        unitsheets.append(sheet)
        print(sheet.title)

    print("\nSheet info")
    print("Verb sheet:", verbsheet.title)
    print("Unit sheets:")
    for sheet in unitsheets:
        print(sheet.title)

    return wb, verbsheet, unitsheets

def parseverbs(verbsheet):
    verbdict = {}  #Dictionary of verbs - key is infinitive, item is the verb variable type.
    print("\nParsing verbs")

    #Insert dummy row for setnences with no verbs
    dummyverbi = osu.verbi()
    dummyverbi.infinitive = "-"
    dummyverbi.englanniksi = "-"
    verbdict[dummyverbi.infinitive] = dummyverbi

    #Now process verb list.
    for row in verbsheet.iter_rows(min_row=2):
        #print(row[0].value, row[1].value)
        currentverbi = osu.verbi()
        currentverbi.infinitive = row[0].value
        currentverbi.englanniksi = row[1].value
        #print("Current verb infinitive is ", currentverbi.infinitive)
        #print("Current verb englanniski is ", currentverbi.englanniksi)
        verbdict[currentverbi.infinitive] = currentverbi
        #print(verbdict[row[0].value], verbdict[row[0].value].infinitive, verbdict[row[0].value].englanniksi)
        #print(verbdict[currentverbi.infinitive].englanniksi)
    #print(verbdict)
    #print("asua returns", verbdict["asua"].englanniksi)
    #print("viedä returns", verbdict["viedä"].englanniksi)
    #for verb in verbdict:
    #    print(verb, "should match", verbdict[verb].englanniksi)
    print("Verb list parsing complete.")
    return verbdict

def parseunits(unitsheets, verbs):
    print("\nStarting unit parsing")
    units = []  #List of units.
    lauset = [] #list of lauset
    errorlist = []
    for unit in unitsheets:
        print("Current unit is:", unit)

        #first get basic info on unit:
        currentunit = osu.unit()
        if unit["A1"].value == "Number of the unit (can include text, if you want)":
            #OpenPyXl is importing numbers as floats - possibly look into this later.
                #But the problem is, of course, that people could number units with floats.  Hm.
            #print(type(unit["B1"].value))
            currentunit.humanid = unit["B1"].value
        else:
            print("WARNING: No unit number detected")
            errorlist.append("Unit parsing error: No unit number detected for unit {}".format(unit))
        if unit["A2"].value == "Unit name":
            currentunit.name = unit["B2"].value
        else:
            print("WARNING: No unit name detected")
            errorlist.append("Unit parsing error: No unit name detected for unit {}".format(unit))
        if unit["A3"].value == "Unit description":
            currentunit.description = unit["B3"].value
        else:
            print("WARNING: No unit description detected")
            errorlist.append("Unit parsing error: No unit description detected for unit {}".format(unit))
        if unit["A4"].value == "Update date":
            currentunit.update_date = unit["B4"].value
        else:
            print("WARNING: No unit update date detected")
            errorlist.append("Unit parsing error: No unit update date detected for unit {}".format(unit))
        if unit["A5"].value == "Other info":
            currentunit.other_info = unit["B5"].value
        else:
            print("WARNING: No unit other info detected")
            errorlist.append("Unit parsing error: No unit other info detected for unit {}".format(unit))
        units.append(currentunit)
        print("Basic info for the unit imported.")

        print("Starting sentence import.")

        #Pull all sentences from the unit
        count = 0
        for row in unit.iter_rows(min_row=8):
            count += 1
            currentlause = osu.lause()
            #print(row[0].value, row[1].value)
            currentlause.unithumanid = currentunit.humanid

            # Column 1 has human ID
            currentlause.humanid = row[0].value
            if currentlause.humanid is None: #No Human-readable ID.  Will crash DB search on attmepting to add it to DB.  So abort.
                print("ERROR: A sentence has no ID number in the spreadsheet - the import of this sentence is being aborted.  Sentence text is {}".format(row[6].value))
                errorlist.append("ERROR: A sentence has no ID number in the spreadsheet - the import of this sentence is being aborted.  Sentence text is {}".format(row[6].value))
                continue #Abort the sentence import

            #Column 2 has type of language flag
            if row[1].value.lower() == "kirjakieli":
                currentlause.kirjakieli = True
            elif row[1].value.lower() == "puhekieli":
                currentlause.puhekieli = True
            else:
                print("WARNING: ei puhekili ta kirjakieli lausessa")
                errorlist.append("WARNING: ei puhekili ta kirjakieli lausessa for lause {}".format(row[3].value))

            #Column 3 has type of thing missing flag
            if row[2].value.lower() == "verb":
                currentlause.type = 1
            elif row[2].value.lower() == "noun":  #Not implemented yet.
                currentlause.type = 2
            elif row[2].value.lower() == "other":
                currentlause.type = 3
            else:
                print("WARNING: 'type of thing missing' text was missing or not recognized for lause {}.  Setting it to 'other' to prevent crashes.  ".format(row[0].value))
                errorlist.append("WARNING: 'type of thing missing' text was missing or not recognized for lause {}  Setting it to 'other' to prevent crashes.  ".format(row[0].value))
                currentlause.type = 3

            if currentlause.type == 1: # it is set to be a verb
                if row[3].value is None: #There is no verb text for a sentence that has listed verb as a type.
                    print("WARNING: Verb text is missing from a sentence that is listed as a verb type - lause {}.  Inserting a - as verb type to avoid crashes".format(row[0].value))
                    errorlist.append("WARNING: Verb text is missing from a sentence that is listed as a verb type - lause {}. Inserting a - as verb type to avoid crashes".format(row[0].value))
                    currentlause.verbi_inf = "-" #prevent crashes by putting in a -
                else: currentlause.verbi_inf = row[3].value
            else:
                currentlause.verbi_inf = "-"   #hard-coding in a dash here to prevent DB errors when coding searches with empty verb field.

            currentlause.hint = row[4].value
            currentlause.vastaus = row[5].value
            if currentlause.vastaus is None:
                print("WARNING: The answer field has been left blank for lause {}".format(row[0].value))
                errorlist.append("WARNING: The answer field has been left blank for lause {}".format(row[0].value))


            #Check if the sentence actually has three ### in it (which also checks if there is even a sentence there!).
            currentlause.lause = row[6].value
            if currentlause.lause is None:
                print("WARNING: there is no sentence text in lause {}".format(row[0].value))
                errorlist.append("WARNING: there is no sentence text in lause {}".format(row[0].value))
            elif re.search('###', currentlause.lause) is None: #check if there are three ###'s, which will be required for later functions.
                print("WARNING: there is no answer position (###) in lause {}".format(row[0].value))
                errorlist.append("WARNING: there is no answer position (###) in lause {}".format(row[0].value))
            elif re.search('####+', currentlause.lause): #check if there are more than three ###'s - no crashes as a result, but uglier formatting.
                print("WARNING: there are more than three #'s in the sentence text for lause {}".format(row[0].value))
                errorlist.append("WARNING: there are more than three #'s in the sentence text for lause {}".format(row[0].value))


            currentlause.lause_englanniksi = row[7].value
            if currentlause.lause_englanniksi is None:
                print("WARNING: The English translation of the sentence has been left blank for lause {}".format(row[0].value))
                errorlist.append("WARNING: The English translation of the sentence has been left blank for lause {}".format(row[0].value))

            # check and see if verb in the unit spreadsheet is in the verb dictionary
            try:
                english = verbs[currentlause.verbi_inf].englanniksi
            except:
                print("WARNING: Verb in the spreadsheet is not in the verb dictionary")
                errorlist.append(
                    "Unit parsing ERROR: No verb definition for a verb {} used in sentence: {}, ".format(currentlause.verbi_inf, currentlause.lause))

            lauset.append(currentlause)

        print("Finishing sentence import for that unit. We imported {} sentences for this unit.".format(count))

    print("\nUnit file parsing finished.  Here are all the units we imported:")
    for unit in units:
        print(unit.name)

    print("\nWe imported {} sentences from the Excel file.".format(len(lauset)))
    #print ("Here are the setnences we imported")
    #for lause in lauset:
    #    print(lause.lause, lause.lause_englanniksi)
    return units, lauset, errorlist


def parsefile(wb, verbsheet, unitsheets):
    verbs = parseverbs(verbsheet)
    units, lauset, errorlist = parseunits(unitsheets, verbs)
    return verbs, units, lauset, errorlist




if __name__ == "__main__":
    print("Executing this script directly, eh?  Try using opinsuomea.py - this isn't supposed to be run on its own.")
    time.sleep(5)
    print("Continuing, but all data will just be discarded at the end")
    wb, verbsheet, unitsheets = openfile()
    verblist, unitlist, lauselist, errorlist = parsefile(wb, verbsheet, unitsheets)
